from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from typing import List, Optional
from app import crud, schemas
from app.database import get_db

router = APIRouter()


@router.get("/", response_model=List[schemas.ObjectResponse])
def get_objects(
    skip: int = 0,
    limit: int = 100,
    pipeline_id: Optional[str] = None,
    object_type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get all objects with optional filtering"""
    objects = crud.get_objects(
        db,
        skip=skip,
        limit=limit,
        pipeline_id=pipeline_id,
        object_type=object_type
    )
    return objects


@router.get("/{object_id}", response_model=schemas.ObjectWithInspections)
def get_object(object_id: int, db: Session = Depends(get_db)):
    """Get a specific object with all its inspections"""
    obj = crud.get_object(db, object_id=object_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Object not found")
    return obj


@router.post("/", response_model=schemas.ObjectResponse)
def create_object(obj: schemas.ObjectCreate, db: Session = Depends(get_db)):
    """Create a new object"""
    # Check if object_id already exists
    existing = crud.get_object(db, object_id=obj.object_id)
    if existing:
        raise HTTPException(status_code=400, detail="Object ID already exists")
    return crud.create_object(db, obj)


@router.get("/map/markers", response_model=List[schemas.ObjectWithInspections])
def get_map_markers(
    method: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    risk_level: Optional[str] = None,
    defect_found: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """Get objects for map visualization with filters"""
    filters = {}
    if method:
        filters['method'] = method
    if date_from:
        filters['date_from'] = date_from
    if date_to:
        filters['date_to'] = date_to
    if risk_level:
        filters['risk_level'] = risk_level
    if defect_found is not None:
        filters['defect_found'] = defect_found
    
    objects = crud.get_objects_with_inspections(db, filters=filters if filters else None)
    return objects


@router.get("/export/excel")
def export_to_excel(
    pipeline_id: Optional[str] = None,
    object_type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export objects and inspections to Excel file"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from datetime import datetime
    
    # Get objects with inspections
    objects = crud.get_objects(
        db,
        skip=0,
        limit=1000,
        pipeline_id=pipeline_id,
        object_type=object_type
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Objects & Inspections"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        "Object ID", "Object Name", "Type", "Pipeline", "Latitude", "Longitude",
        "Year", "Material", "Inspection Date", "Method", "Defect Found",
        "Quality Grade", "ML Risk", "Depth", "Length", "Width"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    row_num = 2
    for obj in objects:
        # Get object with inspections
        obj_full = crud.get_object(db, obj.object_id)
        
        if obj_full.inspections:
            for insp in obj_full.inspections:
                ws.cell(row=row_num, column=1, value=obj.object_id)
                ws.cell(row=row_num, column=2, value=obj.object_name)
                ws.cell(row=row_num, column=3, value=obj.object_type)
                ws.cell(row=row_num, column=4, value=obj.pipeline_id)
                ws.cell(row=row_num, column=5, value=obj.lat)
                ws.cell(row=row_num, column=6, value=obj.lon)
                ws.cell(row=row_num, column=7, value=obj.year)
                ws.cell(row=row_num, column=8, value=obj.material)
                ws.cell(row=row_num, column=9, value=insp.date.strftime("%Y-%m-%d") if insp.date else "")
                ws.cell(row=row_num, column=10, value=insp.method)
                ws.cell(row=row_num, column=11, value="Yes" if insp.defect_found else "No")
                ws.cell(row=row_num, column=12, value=insp.quality_grade or "")
                ws.cell(row=row_num, column=13, value=insp.ml_label or "")
                ws.cell(row=row_num, column=14, value=insp.param1)
                ws.cell(row=row_num, column=15, value=insp.param2)
                ws.cell(row=row_num, column=16, value=insp.param3)
                row_num += 1
        else:
            # Object without inspections
            ws.cell(row=row_num, column=1, value=obj.object_id)
            ws.cell(row=row_num, column=2, value=obj.object_name)
            ws.cell(row=row_num, column=3, value=obj.object_type)
            ws.cell(row=row_num, column=4, value=obj.pipeline_id)
            ws.cell(row=row_num, column=5, value=obj.lat)
            ws.cell(row=row_num, column=6, value=obj.lon)
            ws.cell(row=row_num, column=7, value=obj.year)
            ws.cell(row=row_num, column=8, value=obj.material)
            row_num += 1
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    filename = f"integrityos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

